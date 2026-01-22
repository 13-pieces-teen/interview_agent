"""Export service for generating markdown files from interview experiences."""

from typing import List, Optional, Dict, Any
from datetime import datetime
import io
import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from src.models.schema import InterviewExperience, Question
from src.utils.database import Database
from src.exporters.feishu_exporter import FeishuExporter


class ExportService:
    """Service for exporting interview experiences to markdown."""

    def __init__(self, db: Database):
        """Initialize export service.

        Args:
            db: Database instance
        """
        self.db = db
        self._feishu_exporter: Optional[FeishuExporter] = None

    def get_feishu_exporter(self) -> Optional[FeishuExporter]:
        """Get Feishu exporter instance (lazy initialization).

        Returns:
            FeishuExporter instance if credentials are configured, None otherwise
        """
        if self._feishu_exporter is not None:
            return self._feishu_exporter

        # Try to initialize from environment variables
        app_id = os.getenv("FEISHU_APP_ID")
        app_secret = os.getenv("FEISHU_APP_SECRET")
        api_base = os.getenv("FEISHU_API_BASE", "https://open.feishu.cn/open-apis")

        if app_id and app_secret:
            try:
                self._feishu_exporter = FeishuExporter(
                    app_id=app_id,
                    app_secret=app_secret,
                    api_base=api_base
                )
                return self._feishu_exporter
            except Exception as e:
                print(f"Warning: Failed to initialize Feishu exporter: {e}")
                return None

        return None

    def export_by_interview(
        self,
        experience_ids: Optional[List[str]] = None,
        company_name: Optional[str] = None,
        tags: Optional[List[str]] = None,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        interview_stage: Optional[str] = None,
    ) -> str:
        """Export experiences organized by interview (面经排列).

        Args:
            experience_ids: Optional list of specific experience IDs to export
            company_name: Optional company name filter
            tags: Optional tags filter
            start_date: Optional start date filter (ISO format)
            end_date: Optional end date filter (ISO format)
            interview_stage: Optional interview stage filter

        Returns:
            Markdown formatted string
        """
        # Get experiences
        if experience_ids:
            experiences = [self.db.get_experience(exp_id) for exp_id in experience_ids]
            experiences = [exp for exp in experiences if exp is not None]
        else:
            result = self.db.list_experiences(
                company_name=company_name,
                tags=tags,
                start_date=start_date,
                end_date=end_date,
                interview_stage=interview_stage,
                limit=1000,
            )
            experiences = [self.db.get_experience(exp["id"]) for exp in result]
            experiences = [exp for exp in experiences if exp is not None]

        # Generate markdown
        markdown_lines = ["# 面试记录导出", ""]
        markdown_lines.append(f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        markdown_lines.append(f"总计: {len(experiences)} 条面经")
        markdown_lines.append("")

        for exp in experiences:
            markdown_lines.extend(self._format_experience(exp))
            markdown_lines.append("")

        return "\n".join(markdown_lines)

    def export_by_question(
        self,
        company_name: Optional[str] = None,
        tags: Optional[List[str]] = None,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        interview_stage: Optional[str] = None,
    ) -> str:
        """Export questions grouped by question text (题目排列).

        Args:
            company_name: Optional company name filter
            tags: Optional tags filter
            start_date: Optional start date filter (ISO format)
            end_date: Optional end date filter (ISO format)
            interview_stage: Optional interview stage filter

        Returns:
            Markdown formatted string
        """
        # Get grouped questions
        groups = self.db.get_grouped_questions(
            company_name=company_name,
            tags=tags,
            start_date=start_date,
            end_date=end_date,
            interview_stage=interview_stage,
            limit=1000,
        )

        # Generate markdown
        markdown_lines = ["# 面试题目导出", ""]
        markdown_lines.append(f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        markdown_lines.append(f"总计: {len(groups)} 道题目")
        markdown_lines.append("")

        for idx, group in enumerate(groups, 1):
            markdown_lines.extend(self._format_question_group(idx, group))
            markdown_lines.append("")

        return "\n".join(markdown_lines)

    def _format_experience(self, exp: InterviewExperience) -> List[str]:
        """Format a single experience as markdown.

        Args:
            exp: Interview experience

        Returns:
            List of markdown lines
        """
        lines = []

        # Title
        title_parts = []
        if exp.company_name:
            title_parts.append(exp.company_name)
        if exp.position:
            title_parts.append(exp.position)
        if exp.interview_stage:
            title_parts.append(exp.interview_stage)

        title = " - ".join(title_parts) if title_parts else "未命名面经"
        lines.append(f"## {title}")
        lines.append("")

        # Tags
        if exp.tags:
            tags_str = ", ".join(exp.tags)
            lines.append(f"**标签**: {tags_str}")
            lines.append("")

        # Company info
        if exp.company_scale:
            lines.append(f"**公司规模**: {exp.company_scale}")
        if exp.interview_experience:
            lines.append(f"**面试体验**: {exp.interview_experience}")
        if exp.notes:
            lines.append(f"**备注**: {exp.notes}")

        if exp.company_scale or exp.interview_experience or exp.notes:
            lines.append("")

        # Questions
        if exp.questions:
            lines.append("**问题列表**:")
            lines.append("")
            for idx, q in enumerate(exp.questions, 1):
                lines.append(f"### {idx}. {q.question}")
                lines.append("")

                if q.answer:
                    answer_label = "**答案**"
                    if q.is_ai_generated:
                        answer_label += " (AI生成)"
                    lines.append(f"{answer_label}: {q.answer}")
                    lines.append("")

                # if q.tags:
                #     q_tags_str = ", ".join(q.tags)
                #     lines.append(f"*标签*: {q_tags_str}")
                #     lines.append("")

        # Metadata
        lines.append("---")
        # lines.append(f"*创建时间: {exp.created_at.strftime('%Y-%m-%d %H:%M:%S')}*")
        lines.append("")

        return lines

    def _format_question_group(self, index: int, group: dict) -> List[str]:
        """Format a question group as markdown.

        Args:
            index: Question index
            group: Question group data

        Returns:
            List of markdown lines
        """
        lines = []

        # Question title
        lines.append(f"## {index}. {group['question']}")
        lines.append("")

        # Occurrence count
        count = group.get('count', group.get('occurrence_count', 0))
        lines.append(f"**出现次数**: {count}")
        lines.append("")

        # Tags
        if group.get('tags'):
            tags_str = ", ".join(group['tags'])
            lines.append(f"**标签**: {tags_str}")
            lines.append("")

        # Occurrences
        lines.append("**出现在以下面经**:")
        lines.append("")

        for occ in group['occurrences']:
            # Build occurrence info
            info_parts = []
            if occ.get('company_name'):
                info_parts.append(occ['company_name'])
            if occ.get('position'):
                info_parts.append(occ['position'])
            if occ.get('interview_stage'):
                info_parts.append(occ['interview_stage'])

            info = " - ".join(info_parts) if info_parts else "未命名面经"
            lines.append(f"- {info}")

            # Add answer if available
            if occ.get('answer'):
                answer_label = "答案"
                if occ.get('is_ai_generated'):
                    answer_label += " (AI生成)"
                lines.append(f"  - {answer_label}: {occ['answer']}")

        lines.append("")
        lines.append("---")
        lines.append("")

        return lines

    def export_questions_to_excel(
        self,
        company_name: Optional[str] = None,
        tags: Optional[List[str]] = None,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        interview_stage: Optional[str] = None,
    ) -> bytes:
        """Export questions to Excel format.

        Args:
            company_name: Optional company name filter
            tags: Optional tags filter
            start_date: Optional start date filter (ISO format)
            end_date: Optional end date filter (ISO format)
            interview_stage: Optional interview stage filter

        Returns:
            Excel file content as bytes
        """
        # Get grouped questions
        groups = self.db.get_grouped_questions(
            company_name=company_name,
            tags=tags,
            start_date=start_date,
            end_date=end_date,
            interview_stage=interview_stage,
            limit=1000,
        )

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "面试题目"

        # Define header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Set headers
        headers = ["题目", "出现次数", "标签(多选)", "答案", "公司", "职位"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Set column widths
        ws.column_dimensions['A'].width = 50  # 题目
        ws.column_dimensions['B'].width = 12  # 出现次数
        ws.column_dimensions['C'].width = 25  # 标签
        ws.column_dimensions['D'].width = 60  # 答案
        ws.column_dimensions['E'].width = 20  # 公司
        ws.column_dimensions['F'].width = 25  # 职位

        # Content style
        content_alignment = Alignment(vertical="top", wrap_text=True)

        # Fill data
        current_row = 2
        for group in groups:
            question_text = group['question']
            count = group.get('count', group.get('occurrence_count', 0))
            tags_str = ", ".join(group.get('tags', []))

            # For each occurrence, create a row
            occurrences = group.get('occurrences', [])
            if not occurrences:
                # If no occurrences, create one row with basic info
                ws.cell(row=current_row, column=1, value=question_text).alignment = content_alignment
                ws.cell(row=current_row, column=2, value=count).alignment = content_alignment
                ws.cell(row=current_row, column=3, value=tags_str).alignment = content_alignment
                current_row += 1
            else:
                # Create a row for the first occurrence
                first_occ = occurrences[0]
                ws.cell(row=current_row, column=1, value=question_text).alignment = content_alignment
                ws.cell(row=current_row, column=2, value=count).alignment = content_alignment
                ws.cell(row=current_row, column=3, value=tags_str).alignment = content_alignment
                ws.cell(row=current_row, column=4, value=first_occ.get('answer', '')).alignment = content_alignment
                ws.cell(row=current_row, column=5, value=first_occ.get('company_name', '')).alignment = content_alignment
                ws.cell(row=current_row, column=6, value=first_occ.get('position', '')).alignment = content_alignment
                current_row += 1

        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def export_to_feishu(
        self,
        experience_ids: Optional[List[str]] = None,
        company_name: Optional[str] = None,
        tags: Optional[List[str]] = None,
        folder_token: Optional[str] = None,
        create_index: bool = True
    ) -> Dict[str, Any]:
        """Export experiences to Feishu documents.

        Args:
            experience_ids: Optional list of specific experience IDs to export
            company_name: Optional company name filter
            tags: Optional tags filter
            folder_token: Optional Feishu folder token to place documents in
            create_index: Whether to create an index document (default: True)

        Returns:
            Dictionary with export results

        Raises:
            Exception: If Feishu is not configured or export fails
        """
        feishu = self.get_feishu_exporter()
        if not feishu:
            raise Exception(
                "Feishu exporter is not configured. "
                "Please set FEISHU_APP_ID and FEISHU_APP_SECRET in your .env file."
            )

        # Get experiences
        if experience_ids:
            experiences = [self.db.get_experience(exp_id) for exp_id in experience_ids]
            experiences = [exp for exp in experiences if exp is not None]
        else:
            result = self.db.list_experiences(
                company_name=company_name,
                tags=tags,
                limit=1000,
            )
            experiences = [self.db.get_experience(exp["id"]) for exp in result]
            experiences = [exp for exp in experiences if exp is not None]

        if not experiences:
            return {
                "total": 0,
                "successful": 0,
                "failed": 0,
                "results": [],
                "message": "No experiences found to export"
            }

        # Export to Feishu
        return feishu.export_multiple_to_feishu(
            experiences=experiences,
            folder_token=folder_token,
            create_index=create_index
        )

    def export_single_to_feishu(
        self,
        experience_id: str,
        folder_token: Optional[str] = None
    ) -> Dict[str, str]:
        """Export a single experience to Feishu document.

        Args:
            experience_id: Experience ID to export
            folder_token: Optional Feishu folder token

        Returns:
            Dictionary with document_id, document_url, and title

        Raises:
            Exception: If Feishu is not configured or export fails
        """
        feishu = self.get_feishu_exporter()
        if not feishu:
            raise Exception(
                "Feishu exporter is not configured. "
                "Please set FEISHU_APP_ID and FEISHU_APP_SECRET in your .env file."
            )

        experience = self.db.get_experience(experience_id)
        if not experience:
            raise Exception(f"Experience {experience_id} not found")

        return feishu.export_to_feishu(experience, folder_token)

    def test_feishu_connection(self) -> Dict[str, Any]:
        """Test Feishu API connection.

        Returns:
            Dictionary with connection status

        Raises:
            Exception: If connection test fails
        """
        feishu = self.get_feishu_exporter()
        if not feishu:
            return {
                "configured": False,
                "connected": False,
                "message": "Feishu is not configured. Please set FEISHU_APP_ID and FEISHU_APP_SECRET."
            }

        try:
            feishu.test_connection()
            return {
                "configured": True,
                "connected": True,
                "message": "Successfully connected to Feishu API"
            }
        except Exception as e:
            return {
                "configured": True,
                "connected": False,
                "message": f"Failed to connect: {str(e)}"
            }
